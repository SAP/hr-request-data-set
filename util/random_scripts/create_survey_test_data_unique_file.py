from __future__ import annotations

import glob
import json
from typing import TypedDict

from openpyxl import load_workbook
from tqdm import tqdm


class Ticket(TypedDict):
    ticket: str
    category: str
    sub_category: str


class SurveyTickets(TypedDict):
    age_range: str
    tickets: list[Ticket]


def main() -> None:
    data_path = "ticket_generation/data/survey_tickets"

    files_survey: list[str] = glob.glob(f"{data_path}/Survey_Ticket*.xlsx")

    list_name_sheets: list[str] = [f"Ticket{i}" for i in range(1, 21)]

    tickets_test: list[SurveyTickets] = []

    age_ranges_dict = {
        0: "[0,20]",
        1: "[20,30]",
        2: "[30,40]",
        3: "[40,50]",
        4: "[50,60]",
        5: "60+",
    }

    entities_dict = {
        "Life event_Health issues": ["date_start_absence", "reason", "number_of_days"],
        "Complaint_complaint": ["complaint", "to_who", "reason"],
        "Salary_Gender pay gap": ["wage_gap"],
        "Ask information_Accommodation": ["location", "duration"],
        "Life event_Personal issues": ["description_life_event"],
        "Refund_Refund travel": ["location", "airport", "date_travel"],
        "Salary_Salary raise": ["increase_in_percentage", "work_title", "salary"],
        "Timetable change_Shift change": ["reason_of_change", "date"],
    }

    count_ticket = 0

    for file in tqdm(files_survey):
        wb = load_workbook(file, data_only=True)

        _age_range = "-1"
        initial_sheet_obj = wb["Initial Sheet"]

        assert initial_sheet_obj, "Initial Sheet does not exist"

        for row in range(11, 17):
            _age_check = initial_sheet_obj.cell(row=row, column=2).value
            if _age_check:
                _age_range = age_ranges_dict[row - 11]

        ticket_current_person = []
        for sheet in list_name_sheets:

            sheet_obj = wb[sheet]
            text = sheet_obj.cell(row=21, column=2).value

            # Read the ticket text
            if text:
                count_ticket += 1

                # Read the ticket category and sub category
                category = sheet_obj.cell(row=5, column=2).value
                sub_category = sheet_obj.cell(row=6, column=2).value

                # Read entities
                entities = []
                for _row in range(3, 13):
                    _entity_check = sheet_obj.cell(row=_row, column=3).value
                    if _entity_check:
                        _entity_name = sheet_obj.cell(row=_row, column=1).value
                        entities.append([-1, -1, _entity_name])

                ticket_current_person.append(
                    {
                        "id": count_ticket,
                        "ticket": text,
                        "category": category,
                        "sub_category": sub_category,
                        "entities": entities,
                    }
                )

                with open(
                    f"ticket_generation/data/survey_texts_entities_v2/ticket{count_ticket}.txt", "w", encoding="utf-8"
                ) as f:
                    f.write(text + ";\n")
                    f.write(category + ";\n")
                    f.write(sub_category + ";\n")
                    f.write("ENTITIES:\n")
                    for ent in entities_dict[f"{category}_{sub_category}"]:
                        f.write(ent + ",," + ";\n")

        tickets_test.append({"age_range": _age_range, "tickets": ticket_current_person})

    flat_map = lambda f, xs: (y for ys in xs for y in f(ys))
    tickets_survey_flatted = list(flat_map(lambda t: t["tickets"], tickets_test))

    with open(f"{data_path}/data.json", "w", encoding="utf-8") as f:
        json.dump(tickets_survey_flatted, f, ensure_ascii=False)


if __name__ == "__main__":
    main()
