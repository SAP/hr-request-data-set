from __future__ import annotations

import glob
import json
from typing import TypedDict

from openpyxl import load_workbook
from tqdm import tqdm


class Ticket(TypedDict):
    ticket: str
    category: str
    sub_category: str


class SurveyTickets(TypedDict):
    age_range: str
    tickets: list[Ticket]


def main() -> None:
    path_data = "ticket_generation/data/survey_tickets"

    files_survey: list[str] = glob.glob(f"{path_data}/Survey_Ticket*.xlsx")

    list_name_sheets: list[str] = [f"Ticket{i}" for i in range(1, 21)]

    tickets_test: list[SurveyTickets] = []

    age_ranges_dict = {
        0: "[0,20]",
        1: "[20,30]",
        2: "[30,40]",
        3: "[40,50]",
        4: "[50,60]",
        5: "60+",
    }

    entities_dict = {
        "Life event_Health issues": ["date_start_absence", "reason", "number_of_days"],
        "Complaint_complaint": ["complaint", "to_who", "reason"],
        "Salary_Gender pay gap": ["wage_gap"],
        "Ask information_Accommodation": ["location", "duration"],
        "Life event_Personal issues": ["description_life_event"],
        "Refund_Refund travel": ["location", "airport", "date_travel"],
        "Salary_Salary raise": ["increase_in_percentage", "work_title", "salary"],
        "Timetable change_Shift change": ["reason_of_change", "date"],
    }

    count_ticket = 0

    for file in tqdm(files_survey):
        wb = load_workbook(file, data_only=True)

        _age_range = "-1"
        initial_sheet_obj = wb["Initial Sheet"]

        assert initial_sheet_obj, "Initial Sheet does not exist"

        for row in range(11, 17):
            _age_check = initial_sheet_obj.cell(row=row, column=2).value
            if _age_check:
                _age_range = age_ranges_dict[row - 11]

        ticket_current_person = []
        for sheet in list_name_sheets:

            sheet_obj = wb[sheet]
            text = sheet_obj.cell(row=21, column=2).value

            # Read the ticket text
            if text:
                count_ticket += 1

                # Read the ticket category and sub category
                category = sheet_obj.cell(row=5, column=2).value
                sub_category = sheet_obj.cell(row=6, column=2).value

                # Read entities
                entities = []
                for _row in range(3, 13):
                    _entity_check = sheet_obj.cell(row=_row, column=3).value
                    if _entity_check:
                        _entity_name = sheet_obj.cell(row=_row, column=1).value
                        entities.append([-1, -1, _entity_name])

                ticket_current_person.append(
                    {
                        "id": count_ticket,
                        "ticket": text,
                        "category": category,
                        "sub_category": sub_category,
                        "entities": entities,
                        "label": f"{category}_{sub_category}",
                    }
                )

        tickets_test.append({"age_range": _age_range, "tickets": ticket_current_person})

    flat_map = lambda f, xs: (y for ys in xs for y in f(ys))
    tickets_survey_flatted = list(flat_map(lambda t: t["tickets"], tickets_test))

    path_data = "ticket_generation/data/survey_texts_entities_v3"

    files_survey: list[str] = glob.glob(f"{path_data}/ticket*.txt")

    count = 0

    for file_txt, ticket in zip(files_survey, tickets_survey_flatted):

        count += 1

        ticket_text = ticket["ticket"]

        text_file = ""
        with open(file_txt) as f:
            text_file = f.read()

        entities_list_string = text_file.split("ENTITIES:")[1].strip().replace("\n", "").strip().split(";")

        print(f"Id ticket: {count}")

        entities = []
        for entity_string in entities_list_string:
            split_string = entity_string.split(",")
            if len(split_string) == 3:
                # entities.append([int(split_string[1]), int(split_string[2]), split_string[0]])

                count_new_lines_before = ticket_text[: int(split_string[1])].count("\n")

                _start = int(split_string[1]) - count_new_lines_before
                _end = int(split_string[2]) - count_new_lines_before

                entities.append([_start, _end, split_string[0]])

                print(
                    f"{split_string[0]}: {ticket_text[int(split_string[1]) - count_new_lines_before: int(split_string[2]) - count_new_lines_before]}"
                )

        ticket["entities"] = entities

        print(f"{'-'*40}")

    with open(f"{path_data}/data_new.json", "w", encoding="utf-8") as f:
        json.dump(tickets_survey_flatted, f, ensure_ascii=False)


if __name__ == "__main__":
    main()
